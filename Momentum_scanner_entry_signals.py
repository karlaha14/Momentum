#!/usr/bin/env python3
"""
Entry Signal Scanner — Condiciones C1 + C3
--------------------------------------------
Lee los tickers de la hoja "Cumplen C6" del archivo Momentum_scanner_C6.xlsx
y evalúa:

  C1: Precio (columna del Excel) <= SMA50 actual
  C3: SMA50 actual > SMA50 de hace 20 días hábiles

SMA50 actual       = promedio de los últimos 50 cierres disponibles (hasta hoy)
SMA50 hace 20 días = promedio de los 50 cierres terminando 20 días hábiles atrás

Los que cumplan AMBAS condiciones se exportan en la hoja "Señal Activa".
Los que no cumplan se exportan en la hoja "Sin Señal".
"""

from urllib.request import urlopen
import urllib.error
import time
import certifi
import json
import os
import pandas as pd
import numpy as np
from datetime import date
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────

API_KEY = "gulW74E75O2BjtTbmERXOInaTOD5BltA"

INPUT_EXCEL  = r"C:\Users\Karla Armas\OneDrive\Desktop\ZONA TRADING\Momentum\Momentum_scanner_C6.xlsx"
OUTPUT_EXCEL = r"C:\Users\Karla Armas\OneDrive\Desktop\ZONA TRADING\Momentum\Momentum_entry_signals.xlsx"

API_CALL_DELAY = 0.5   # segundos entre llamadas — subir a 1.0 si hay error 429

# ─── FECHAS ───────────────────────────────────────────────────────────────────

today = date.today()

# Necesitamos ~90 días hábiles de historia para cubrir:
# 50 días (SMA50 actual) + 20 días hacia atrás + buffer
fetch_from = today - relativedelta(days=130)

# ─── API ──────────────────────────────────────────────────────────────────────

def get_json(url: str, retries: int = 3):
    for attempt in range(retries):
        try:
            time.sleep(API_CALL_DELAY)
            response = urlopen(url, cafile=certifi.where())
            return json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            if e.code == 429:
                wait = 10 * (attempt + 1)
                print(f"    ⏳ Rate limit (429) — esperando {wait}s...")
                time.sleep(wait)
            else:
                raise
    raise RuntimeError(f"Fallo tras {retries} intentos (429)")

# ─── FUNCIONES ────────────────────────────────────────────────────────────────

def fetch_prices(ticker: str) -> pd.DataFrame:
    """Descarga precios diarios EOD desde fetch_from hasta hoy."""
    url = (
        f"https://financialmodelingprep.com/stable/historical-price-eod/light"
        f"?symbol={ticker}&from={fetch_from}&to={today}&apikey={API_KEY}"
    )
    raw = get_json(url)
    if not raw:
        return pd.DataFrame()
    df = pd.DataFrame(raw)
    df["date"] = pd.to_datetime(df["date"])
    df = df[["date", "price"]].rename(columns={"price": "close"})
    return df.sort_values("date").reset_index(drop=True)


def calc_sma50_current(closes: pd.Series) -> float:
    """SMA50 usando los últimos 50 cierres disponibles."""
    if len(closes) < 50:
        return np.nan
    return closes.iloc[-50:].mean()


def calc_sma50_20days_ago(closes: pd.Series) -> float:
    """
    SMA50 de hace 20 días hábiles:
    toma los 50 cierres que terminaban en la posición [-20] del array.
    Es decir, la ventana va de [-70] a [-21] (50 valores).
    """
    needed = 50 + 20  # 70 días hábiles mínimo
    if len(closes) < needed:
        return np.nan
    # El cierre de hace 20 días hábiles está en índice -20 (no incluido en slice)
    window = closes.iloc[-70:-20]
    if len(window) < 50:
        return np.nan
    return window.mean()


def evaluate_ticker(ticker: str, price_from_excel: float) -> dict:
    """Descarga precios y evalúa C1 y C3 para un ticker."""
    df = fetch_prices(ticker)

    base = {
        "Ticker"           : ticker,
        "Precio"           : price_from_excel,
        "SMA50 actual"     : np.nan,
        "SMA50 hace 20d"   : np.nan,
        "C1 (P<=SMA50)"    : "FALSE",
        "C3 (SMA50>SMA50-20d)": "FALSE",
        "Señal"            : "Sin señal",
    }

    if df.empty or len(df) < 70:
        base["Señal"] = "Sin datos"
        return base

    closes = df["close"]

    sma50_now = calc_sma50_current(closes)
    sma50_20d = calc_sma50_20days_ago(closes)

    if np.isnan(sma50_now) or np.isnan(sma50_20d):
        base["Señal"] = "Cálculo incompleto"
        return base

    c1 = price_from_excel <= sma50_now
    c3 = sma50_now > sma50_20d

    return {
        "Ticker"               : ticker,
        "Precio"               : round(price_from_excel, 2),
        "SMA50 actual"         : round(sma50_now, 4),
        "SMA50 hace 20d"       : round(sma50_20d, 4),
        "C1 (P<=SMA50)"        : "TRUE" if c1 else "FALSE",
        "C3 (SMA50>SMA50-20d)" : "TRUE" if c3 else "FALSE",
        "Señal"                : "ENTRADA" if (c1 and c3) else "Sin señal",
    }

# ─── LEER INPUT EXCEL ─────────────────────────────────────────────────────────

print("=" * 60)
print("  ENTRY SIGNAL SCANNER — C1 + C3")
print("=" * 60)
print(f"  Fecha de ejecución : {today}")
print(f"  Leyendo            : {INPUT_EXCEL}")
print("=" * 60)

try:
    df_input = pd.read_excel(INPUT_EXCEL, sheet_name="Cumplen C6")
except FileNotFoundError:
    print(f"\n❌ No se encontró el archivo: {INPUT_EXCEL}")
    exit()
except Exception as e:
    print(f"\n❌ Error leyendo Excel: {e}")
    exit()

# Validar columnas requeridas
required = ["Ticker", "Precio"]
missing = [c for c in required if c not in df_input.columns]
if missing:
    print(f"\n❌ Columnas faltantes en 'Cumplen C6': {missing}")
    print(f"   Columnas disponibles: {list(df_input.columns)}")
    exit()

# Limpiar filas sin ticker (pueden venir del encabezado decorativo del Excel)
df_input = df_input.dropna(subset=["Ticker"]).reset_index(drop=True)
tickers_input = list(zip(df_input["Ticker"].astype(str), df_input["Precio"]))

print(f"  Tickers leídos de 'Cumplen C6': {len(tickers_input)}\n")

# ─── EVALUAR CONDICIONES ──────────────────────────────────────────────────────

results = []

for ticker, precio in tickers_input:
    print(f"  [{ticker}]", end=" ", flush=True)
    result = evaluate_ticker(ticker, float(precio) if not pd.isna(precio) else np.nan)
    results.append(result)
    c1 = result["C1 (P<=SMA50)"]
    c3 = result["C3 (SMA50>SMA50-20d)"]
    print(f"SMA50={result['SMA50 actual']}  SMA50-20d={result['SMA50 hace 20d']}  C1={c1}  C3={c3}  → {result['Señal']}")

# ─── SEPARAR RESULTADOS ───────────────────────────────────────────────────────

df_all     = pd.DataFrame(results)
df_active  = df_all[df_all["Señal"] == "ENTRADA"].drop(columns=["Señal"]).reset_index(drop=True)
df_inactive = df_all[df_all["Señal"] != "ENTRADA"].drop(columns=["Señal"]).reset_index(drop=True)
df_active.index  += 1
df_inactive.index += 1

print(f"\n  ✅ Con señal activa (C1 y C3): {len(df_active)}")
print(f"  ⚪ Sin señal                  : {len(df_inactive)}")

# ─── EXPORTAR A EXCEL ─────────────────────────────────────────────────────────

header_note = f"Generado: {today}  |  Evaluados: {len(tickers_input)} tickers de 'Cumplen C6'"


def write_sheet(ws, df: pd.DataFrame, title: str, header_color: str):
    ws.sheet_view.showGridLines = False

    ws.append([title])
    ws.cell(row=ws.max_row, column=1).font = Font(
        bold=True, name="Arial", size=12, color="1F3864"
    )
    ws.append([header_note])
    ws.cell(row=ws.max_row, column=1).font = Font(
        name="Arial", size=9, italic=True, color="666666"
    )
    ws.append([])

    if df.empty:
        ws.append(["Sin registros"])
        return

    cols = list(df.columns)
    header_row = ws.max_row + 1
    ws.append(cols)

    fill   = PatternFill("solid", start_color=header_color, end_color=header_color)
    b_side = Side(style="thin", color="CCCCCC")
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=b_side)

    for i, row_data in enumerate(df.itertuples(index=False), start=1):
        ws.append(list(row_data))
        alt_fill = PatternFill("solid",
                               start_color="F2F2F2" if i % 2 == 0 else "FFFFFF",
                               end_color="F2F2F2" if i % 2 == 0 else "FFFFFF")
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.fill = alt_fill
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=Side(style="thin", color="DDDDDD"))

    for col_idx, col_name in enumerate(cols, start=1):
        col_vals = [len(str(v)) for v in df.iloc[:, col_idx - 1]] if not df.empty else []
        max_len  = max([len(str(col_name))] + col_vals)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)


os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)
wb = Workbook()

ws1 = wb.active
ws1.title = "Señal Activa"
write_sheet(ws1, df_active, "SEÑAL DE ENTRADA ACTIVA — C1 y C3 cumplidas", "1A7A4A")

ws2 = wb.create_sheet("Sin Señal")
write_sheet(ws2, df_inactive, "SIN SEÑAL — C1 o C3 no cumplidas", "555555")

wb.save(OUTPUT_EXCEL)

print(f"\n  ✅ Excel guardado en: {OUTPUT_EXCEL}")
print("=" * 60)
print(f"  Fecha             : {today}")
print(f"  Tickers evaluados : {len(tickers_input)}")
print(f"  Señal activa      : {len(df_active)}")
print(f"  Sin señal         : {len(df_inactive)}")
print("=" * 60)