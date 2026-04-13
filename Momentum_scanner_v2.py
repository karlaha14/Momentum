#!/usr/bin/env python3
"""
Momentum Scanner — Condición C6
---------------------------------
1. Calcula el Score bruto de momentum para cada ticker:
   Score bruto = R_12,1 / σ
   donde R_12,1 = retorno simple de 12 meses excluyendo el último mes
         σ      = desviación estándar de retornos diarios en ese período

2. Toma el Top 100 por Score bruto.

3. Aplica condición C6:
   (Último Precio Close / Precio máximo 52 semanas) >= 0.60

4. Exporta a Excel en dos hojas:
   - Hoja 1: tickers que cumplen C6
   - Hoja 2: tickers que NO cumplen C6
"""

from urllib.request import urlopen
import urllib.error
import time
import certifi
import json
import os
import pandas as pd
import numpy as np
from datetime import date
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────

API_KEY = "gulW74E75O2BjtTbmERXOInaTOD5BltA"

EXCEL_PATH = r"C:\Users\Karla Armas\OneDrive\Desktop\ZONA TRADING\Momentum\Momentum_scanner_C6.xlsx"

API_CALL_DELAY = 1.0   # segundos entre llamadas — subir a 1.0 si hay error 429

TICKERS = [
        "A","AA","AAL","AAON","AAPL","ABBV","ABNB","ABT","ACGL","ACHC","ACI","ACM","ACN","ADBE","ADC","ADI","ADM","ADP","ADSK","ADT","AEE",
    "AEP","AES","AFG","AFL","AFRM","AGCO","AGNC","AGO","AIG","AIT","AIZ","AJG","AKAM","AL","ALAB","ALB","ALGM","ALGN","ALK","ALL","ALLE",
    "ALLY","ALNY","ALSN","AM","AMAT","AMCR","AMD","AME","AMG","AMGN","AMH","AMKR","AMP","AMT","AMTM","AMZN","AN","ANET","AON","AOS","APA",
    "APD","APG","APH","APLS","APO","APP","APPF","APTV","AR","ARE","ARES","ARMK","ARW","AS","ASH","ASTS","ATI","ATO","ATR","AU","AUR","AVB",
    "AVGO","AVT","AVTR","AVY","AWI","AWK","AXON","AXP","AXS","AXTA","AYI","AZO","BA","BAC","BAH","BALL","BAM","BAX","BBWI","BBY","BC","BDX",
    "BEN","BEPC","BF-A","BFAM","BF-B","BG","BHF","BIIB","BILL","BIO","BIRK","BJ","BK","BKNG","BKR","BLD","BLDR","BLK","BMRN","BMY","BOKF",
    "BPOP","BR","BRBR","BRK-B","BRKR","BRO","BROS","BRX","BSX","BSY","BURL","BWA","BWXT","BX","BXP","BYD","C","CACC","CACI","CAG","CAH",
    "CAR","CARR","CART","CASY","CAT","CAVA","CB","CBOE","CBRE","CBSH","CCCS","CCI","CCK","CCL","CDNS","CDW","CE","CEG","CELH","CERT","CF",
    "CFG","CFLT","CFR","CG","CGNX","CHD","CHDN","CHE","CHH","CHRD","CHRW","CHTR","CHWY","CI","CIEN","CINF","CIVI","CL","CLF","CLH","CLVT",
    "CLX","CMA","CMCSA","CME","CMG","CMI","CMS","CNA","CNC","CNH","CNM","CNP","CNXC","COF","COHR","COIN","COKE","COLB","COLD","COLM","COO",
    "COOP","COP","COR","CORT","COST","COTY","CPAY","CPB","CPNG","CPRT","CPT","CR","CRH","CRL","CRM","CROX","CRS","CRUS","CRWD","CSCO","CSGP",
    "CSL","CSX","CTAS","CTRA","CTSH","CTVA","CUBE","CUZ","CVNA","CVS","CVX","CW","CWEN","CWE-A","CXT","CZR","D","DAL","DAR","DASH","DAY","DBX",
    "DCI","DD","DDOG","DDS","DE","DECK","DELL","DG","DGX","DHI","DHR","DINO","DIS","DJT","DKNG","DKS","DLB","DLR","DLTR","DNB","DOC","DOCS",
    "DOCU","DOV","DOW","DOX","DPZ","DRI","DRS","DT","DTE","DTM","DUK","DUOL","DV","DVA","DVN","DXC","DXCM","EA","EBAY","ECG","ECL","ED",
    "EEFT","EFX","EG","EGP","EHC","EIX","EL","ELAN","ELF","ELS","ELV","EME","EMN","EMR","ENPH","ENTG","EOG","EPAM","EPR","EQH","EQIX","EQR",
    "EQT","ES","ESAB","ESI","ESS","ESTC","ETN","ETR","ETSY","EVR","EVRG","EW","EWBC","EXAS","EXC","EXE","EXEL","EXLS","EXP","EXPD","EXPE",
    "EXR","F","FAF","FANG","FAST","FBIN","FCN","FCNCA","FCX","FDS","FDX","FE","FERG","FFIV","FHB","FHN","FI","FICO","FIS","FITB","FIVE",
    "FIX","FLEX","FLO","FLS","FLUT","FMC","FNB","FND","FNF","FOUR","FOX","FOXA","FR","FRHC","FRPT","FRT","FSLR","FTAI","FTI","FTNT","FTV",
    "FWONA","FWONK","FYBR","G","GAP","GD","GDDY","GE","GEHC","GEN","GEV","GFS","GGG","GILD","GIS","GL","GLIBA","GLIBK","GLOB","GLPI","GLW",
    "GM","GME","GMED","GNRC","GNTX","GOOG","GOOGL","GPC","GPK","GPN","GRMN","GS","GTES","GTLB","GTM","GWRE","GWW","GXO","H","HAL","HALO",
    "HAS","HAYW","HBAN","HCA","HD","HEI","HEI-A","HHH","HIG","HII","HIW","HLI","HLNE","HLT","HOG","HOLX","HON","HOOD","HPE","HPQ","HR","HRB",
    "HRL","HSIC","HST","HSY","HUBB","HUBS","HUM","HUN","HWM","HXL","IAC","IBKR","IBM","ICE","IDA","IDXX","IEX","IFF","ILMN","INCY","INFA",
    "INGM","INGR","INSM","INSP","INTC","INTU","INVH","IONS","IOT","IP","IPG","IPGP","IQV","IR","IRDM","IRM","ISRG","IT","ITT","ITW","IVZ",
    "J","JAZZ","JBHT","JBL","JCI","JEF","JHG","JHX","JKHY","JLL","JNJ","JPM","K","KBR","KD","KDP","KEX","KEY","KEYS","KHC","KIM","KKR","KLAC",
    "KMB","KMI","KMPR","KMX","KNSL","KNX","KO","KR","KRC","KRMN","KVUE","L","LAD","LAMR","LAZ","LBRDA","LBRDK","LBTYA","LBTYK","LCID","LDOS",
    "LEA","LECO","LEN","LEN-B","LFUS","LH","LHX","LII","LIN","LINE","LITE","LKQ","LLY","LLYVA","LLYVK","LMT","LNC","LNG","LNT","LNW","LOAR",
    "LOPE","LOW","LPLA","LPX","LRCX","LSCC","LSTR","LULU","LUV","LVS","LW","LYB","LYFT","LYV","M","MA","MAA","MAN","MANH","MAR","MAS","MASI",
    "MAT","MCD","MCHP","MCK","MCO","MDB","MDLZ","MDT","MDU","MEDP","MET","META","MGM","MHK","MIDD","MKC","MKL","MKSI","MKTX","MLI","MLM","MMC",
    "MMM","MNST","MO","MOH","MORN","MOS","MP","MPC","MPW","MPWR","MRK","MRNA","MRP","MRVL","MS","MSA","MSCI","MSFT","MSGS","MSI","MSM","MSTR",
    "MTB","MTCH","MTD","MTDR","MTG","MTN","MTSI","MTZ","MU","MUSA","NBIX","NCLH","NCNO","NDAQ","NDSN","NEE","NEM","NET","NEU","NFG","NFLX","NI",
    "NKE","NLY","NNN","NOC","NOV","NOW","NRG","NSA","NSC","NTAP","NTNX","NTRA","NTRS","NU","NUE","NVDA","NVR","NVST","NVT","NWL","NWS","NWSA",
    "NXST","NYT","O","OC","ODFL","OGE","OGN","OHI","OKE","OKTA","OLED","OLLI","OLN","OMC","OMF","ON","ONON","ONTO","ORCL","ORI","ORLY","OSK",
    "OTIS","OVV","OWL","OXY","OZK","PAG","PANW","PATH","PAYC","PAYX","PB","PCAR","PCG","PCOR","PCTY","PEG","PEGA","PEN","PENN","PEP","PFE","PFG",
    "PFGC","PG","PGR","PH","PHM","PINS","PK","PKG","PLD","PLNT","PLTR","PM","PNC","PNFP","PNR","PNW","PODD","POOL","POST","PPC","PPG","PPL",
    "PR","PRGO","PRI","PRMB","PRU","PSA","PSN","PSTG","PSX","PTC","PVH","PWR","PYPL","QCOM","QGEN","QRVO","QS","QSR","QXO","R","RAL","RARE",
    "RBA","RBC","RBLX","RBRK","RCL","RDDT","REG","REGN","REXR","REYN","RF","RGA","RGEN","RGLD","RH","RHI","RITM","RIVN","RJF","RKLB","RKT",
    "RL","RLI","RMD","RNG","RNR","ROIV","ROK","ROKU","ROL","ROP","ROST","RPM","RPRX","RRC","RRX","RS","RSG","RTX","RVMD","RVTY","RYAN","RYN",
    "S","SAIA","SAIC","SAIL","SAM","SARO","SBAC","SBUX","SCCO","SCHW","SCI","SEB","SEE","SEIC","SF","SFD","SFM","SGI","SHC","SHW","SIRI","SITE",
    "SJM","SKX","SLB","SLGN","SLM","SMCI","SMG","SMMT","SN","SNA","SNDK","SNDR","SNOW","SNPS","SNV","SNX","SO","SOFI","SOLV","SON","SPG","SPGI",
    "SPOT","SPR","SRE","SRPT","SSB","SSD","SSNC","ST","STAG","STE","STLD","STT","STWD","STZ","SUI","SW","SWK","SWKS","SYF","SYK","SYY","T","TAP",
    "TDC","TDG","TDY","TEAM","TECH","TEM","TER","TFC","TFSL","TFX","TGT","THC","THG","THO","TIGO","TJX","TKO","TKR","TLN","TMO","TMUS","TNL","TOL",
    "TOST","TPG","TPL","TPR","TREX","TRGP","TRMB","TROW","TRU","TRV","TSCO","TSLA","TSN","TT","TTC","TTD","TTEK","TTWO","TW","TWLO","TXN","TXRH",
    "TXT","TYL","U","UA","UAA","UAL","UBER","UDR","UGI","UHAL","UHA-B","UHS","UI","ULTA","UNH","UNM","UNP","UPS","URI","USB","USFD","UTHR","UWMC",
    "V","VEEV","VFC","VICI","VIK","VIRT","VKTX","VLO","VLTO","VMC","VMI","VNO","VNOM","VNT","VOYA","VRSK","VRSN","VRT","VRTX","VST","VTR","VTRS",
    "VVV","VZ","W","WAB","WAL","WAT","WBA","WBD","WBS","WCC","WDAY","WDC","WEC","WELL","WEN","WEX","WFC","WFRD","WH","WHR","WING","WLK","WM","WMB",
    "WMS","WMT","WPC","WRB","WSC","WSM","WSO","WST","WTFC","WTM","WTRG","WTW","WU","WWD","WY","WYNN","XEL","XOM","XP","XPO","XRAY","XYL","XYZ","ZTS",
    "TEL","NXPI","YUM","STX","ZBH","ERIE","ZBRA","PSKY"
]

# ─── FECHAS DINÁMICAS ─────────────────────────────────────────────────────────

today      = date.today()
end_date   = today.replace(day=1) - relativedelta(days=1)
start_date = end_date - relativedelta(months=12)
fetch_from = start_date - relativedelta(days=7)

print("=" * 60)
print("  MOMENTUM SCANNER — CONDICIÓN C6")
print("=" * 60)
print(f"  Fecha de ejecución  : {today}")
print(f"  Período de momentum : {start_date} → {end_date}")
print("=" * 60)

# ─── API ──────────────────────────────────────────────────────────────────────

def get_json(url: str, retries: int = 3):
    for attempt in range(retries):
        try:
            time.sleep(API_CALL_DELAY)
            response = urlopen(url, cafile=certifi.where())
            return json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            if e.code == 429:
                wait = 10 * (attempt + 1)
                print(f"    ⏳ Rate limit (429) — esperando {wait}s...")
                time.sleep(wait)
            else:
                raise
    raise RuntimeError(f"Fallo tras {retries} intentos (429)")

# ─── FUNCIONES ────────────────────────────────────────────────────────────────

def fetch_prices(ticker: str) -> pd.DataFrame:
    url = (
        f"https://financialmodelingprep.com/stable/historical-price-eod/light"
        f"?symbol={ticker}&from={fetch_from}&to={end_date}&apikey={API_KEY}"
    )
    raw = get_json(url)
    if not raw:
        return pd.DataFrame()
    df = pd.DataFrame(raw)
    df["date"] = pd.to_datetime(df["date"])
    df = df[["date", "price"]].rename(columns={"price": "close"})
    return df.sort_values("date").reset_index(drop=True)


def fetch_quote(ticker: str) -> dict:
    """Obtiene precio actual (price) y máximo 52 semanas (yearHigh) desde /quote."""
    url = (
        f"https://financialmodelingprep.com/stable/quote"
        f"?symbol={ticker}&apikey={API_KEY}"
    )
    raw = get_json(url)
    if not raw:
        return {}
    record = raw[0] if isinstance(raw, list) else raw
    return {
        "price"   : record.get("price", np.nan),
        "yearHigh": record.get("yearHigh", np.nan),
    }


def get_monthly_close(df: pd.DataFrame, target_date: date) -> float:
    mask = df["date"] <= pd.Timestamp(target_date)
    subset = df[mask]
    if subset.empty:
        return np.nan
    return subset.iloc[-1]["close"]

# ─── PASO 1: CALCULAR SCORE BRUTO PARA TODOS LOS TICKERS ─────────────────────

results = []

for ticker in TICKERS:
    print(f"  [{ticker}]", end=" ", flush=True)

    df_prices = fetch_prices(ticker)
    if df_prices.empty:
        print("sin datos")
        continue

    price_t12 = get_monthly_close(df_prices, start_date)
    price_t1  = get_monthly_close(df_prices, end_date)

    if np.isnan(price_t12) or np.isnan(price_t1) or price_t12 == 0:
        print("precios insuficientes")
        continue

    R_12_1 = (price_t1 / price_t12) - 1

    period_mask = (
        (df_prices["date"] >= pd.Timestamp(start_date)) &
        (df_prices["date"] <= pd.Timestamp(end_date))
    )
    df_period = df_prices[period_mask].copy()

    if len(df_period) < 30:
        print("pocos días hábiles")
        continue

    df_period["ret"] = df_period["close"].pct_change()
    sigma = df_period["ret"].std()

    if sigma == 0 or np.isnan(sigma):
        print("volatilidad inválida")
        continue

    raw_score = R_12_1 / sigma
    print(f"Score bruto = {raw_score:.4f}")

    results.append({"Ticker": ticker, "Score_bruto": round(raw_score, 4)})

if not results:
    print("\n❌ Sin resultados. Verifica API key y tickers.")
    exit()

# ─── PASO 2: TOP 100 POR SCORE BRUTO ─────────────────────────────────────────

df_scores = (
    pd.DataFrame(results)
    .sort_values("Score_bruto", ascending=False)
    .head(100)
    .reset_index(drop=True)
)
df_scores.index += 1

print(f"\n  Top 100 seleccionados. Aplicando condición C6...\n")

# ─── PASO 3: APLICAR CONDICIÓN C6 ────────────────────────────────────────────
# C6: (Último Precio Close / yearHigh 52 semanas) >= 0.60

c6_results = []

for _, row in df_scores.iterrows():
    ticker = row["Ticker"]
    score  = row["Score_bruto"]

    print(f"  C6 → [{ticker}]", end=" ", flush=True)

    quote = fetch_quote(ticker)
    price    = quote.get("price", np.nan)
    year_high = quote.get("yearHigh", np.nan)

    if np.isnan(price) or np.isnan(year_high) or year_high == 0:
        ratio = np.nan
        c6    = False
        print("datos insuficientes")
    else:
        ratio = round(price / year_high, 4)
        c6    = ratio >= 0.60
        print(f"Price={price:.2f}  YearHigh={year_high:.2f}  Ratio={ratio:.4f}  C6={'TRUE' if c6 else 'FALSE'}")

    c6_results.append({
        "Ticker"      : ticker,
        "Score_bruto" : score,
        "Precio"      : round(price, 2) if not np.isnan(price) else np.nan,
        "Year High"   : round(year_high, 2) if not np.isnan(year_high) else np.nan,
        "P/YearHigh"  : ratio,
        "C6 Result"   : "TRUE" if c6 else "FALSE",
    })

df_c6 = pd.DataFrame(c6_results)
df_pass = df_c6[df_c6["C6 Result"] == "TRUE"].reset_index(drop=True)
df_fail = df_c6[df_c6["C6 Result"] == "FALSE"].reset_index(drop=True)
df_pass.index += 1
df_fail.index += 1

print(f"\n  ✅ Cumplen C6 : {len(df_pass)}")
print(f"  ❌ No cumplen : {len(df_fail)}")

# ─── PASO 4: EXPORTAR A EXCEL ─────────────────────────────────────────────────

header_note = f"Generado: {today}  |  Período: {start_date} → {end_date}"

def write_sheet(ws, df: pd.DataFrame, title: str, header_color: str):
    ws.sheet_view.showGridLines = False

    # Título
    ws.append([title])
    tc = ws.cell(row=ws.max_row, column=1)
    tc.font = Font(bold=True, name="Arial", size=12, color="1F3864")

    # Nota de período
    ws.append([header_note])
    nc = ws.cell(row=ws.max_row, column=1)
    nc.font = Font(name="Arial", size=9, italic=True, color="666666")

    ws.append([])  # fila vacía

    if df.empty:
        ws.append(["Sin registros"])
        return

    cols = list(df.columns)
    header_row = ws.max_row + 1
    ws.append(cols)

    # Estilo encabezado
    fill   = PatternFill("solid", start_color=header_color, end_color=header_color)
    b_side = Side(style="thin", color="CCCCCC")
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=b_side)

    # Datos con filas alternadas
    for i, row_data in enumerate(df.itertuples(index=False), start=1):
        ws.append(list(row_data))
        row_fill = PatternFill("solid",
                               start_color="F2F2F2" if i % 2 == 0 else "FFFFFF",
                               end_color="F2F2F2" if i % 2 == 0 else "FFFFFF")
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.fill = row_fill
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=Side(style="thin", color="DDDDDD"))

    # Ancho automático
    for col_idx, col_name in enumerate(cols, start=1):
        col_vals = [len(str(v)) for v in df.iloc[:, col_idx - 1]] if not df.empty else []
        max_len  = max([len(str(col_name))] + col_vals)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)


os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
wb = Workbook()

ws1 = wb.active
ws1.title = "Cumplen C6"
write_sheet(ws1, df_pass, "TICKERS QUE CUMPLEN C6  (P / YearHigh >= 0.60)", "1A7A4A")

ws2 = wb.create_sheet("No Cumplen C6")
write_sheet(ws2, df_fail, "TICKERS QUE NO CUMPLEN C6  (P / YearHigh < 0.60)", "7A1A1A")

wb.save(EXCEL_PATH)

print(f"\n  ✅ Excel guardado en: {EXCEL_PATH}")
print("=" * 60)
print(f"  Fecha de análisis : {today}")
print(f"  Período           : {start_date} → {end_date}")
print(f"  Tickers evaluados : {len(TICKERS)}")
print(f"  Top 100 (Score)   : {len(df_scores)}")
print(f"  Cumplen C6        : {len(df_pass)}")
print(f"  No cumplen C6     : {len(df_fail)}")
print("=" * 60)