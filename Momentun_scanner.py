#!/usr/bin/env python3
"""
SPMO Momentum Score Calculator + Entry Signal Detection
---------------------------------------------------------
PARTE 1 — Momentum Score (metodología S&P 500 Momentum Index):
  - Retorno simple de 12 meses excluyendo el último mes (R_12,1)
  - Volatilidad = desviación estándar de retornos diarios en ese período
  - Momentum Score = R_12,1 / σ  (estandarizado como z-score)
  - Peso final = Market Cap × Momentum Score (normalizado)

PARTE 2 — Señal de entrada (Top 50 por Score bruto):
  - SMA 50: media simple de los últimos 50 cierres
  - ATR 15: Average True Range de 15 días (RMA/Wilder)
  - Señal ACTIVA si:
      Condición 1: Precio actual <= SMA 50
      Condición 2: Precio actual - ATR(15) <= SMA 50
"""

from urllib.request import urlopen
import time
import certifi
import json
import os
import pandas as pd
import numpy as np
from datetime import date
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────

API_KEY = "gulW74E75O2BjtTbmERXOInaTOD5BltA"

# Ruta de exportación Excel
EXCEL_PATH = r"C:\Users\HP\OneDrive - Zona Trading\Documents\Momentum\Momentum_scanner_signals.xlsx"

# Pausa entre llamadas a la API (segundos) — ajustar según límite del plan
API_CALL_DELAY = 0.5   # 0.5s → ~120 llamadas/min. Subir a 1.0 si sigue con 429

# Lista de tickers a analizar (reemplaza con tu variable TICKERS)
TICKERS = [
        "A","AA","AAL","AAON","AAPL","ABBV","ABNB","ABT","ACGL","ACHC","ACI","ACM","ACN","ADBE","ADC","ADI","ADM","ADP","ADSK","ADT","AEE",
    "AEP","AES","AFG","AFL","AFRM","AGCO","AGNC","AGO","AIG","AIT","AIZ","AJG","AKAM","AL","ALAB","ALB","ALGM","ALGN","ALK","ALL","ALLE",
    "ALLY","ALNY","ALSN","AM","AMAT","AMCR","AMD","AME","AMG","AMGN","AMH","AMKR","AMP","AMT","AMTM","AMZN","AN","ANET","AON","AOS","APA",
    "APD","APG","APH","APLS","APO","APP","APPF","APTV","AR","ARE","ARES","ARMK","ARW","AS","ASH","ASTS","ATI","ATO","ATR","AU","AUR","AVB",
    "AVGO","AVT","AVTR","AVY","AWI","AWK","AXON","AXP","AXS","AXTA","AYI","AZO","BA","BAC","BAH","BALL","BAM","BAX","BBWI","BBY","BC","BDX",
    "BEN","BEPC","BF-A","BFAM","BF-B","BG","BHF","BIIB","BILL","BIO","BIRK","BJ","BK","BKNG","BKR","BLD","BLDR","BLK","BMRN","BMY","BOKF",
    "BPOP","BR","BRBR","BRK-B","BRKR","BRO","BROS","BRX","BSX","BSY","BURL","BWA","BWXT","BX","BXP","BYD","C","CACC","CACI","CAG","CAH",
    "CAR","CARR","CART","CASY","CAT","CAVA","CB","CBOE","CBRE","CBSH","CCCS","CCI","CCK","CCL","CDNS","CDW","CE","CEG","CELH","CERT","CF",
    "CFG","CFLT","CFR","CG","CGNX","CHD","CHDN","CHE","CHH","CHRD","CHRW","CHTR","CHWY","CI","CIEN","CINF","CIVI","CL","CLF","CLH","CLVT",
    "CLX","CMA","CMCSA","CME","CMG","CMI","CMS","CNA","CNC","CNH","CNM","CNP","CNXC","COF","COHR","COIN","COKE","COLB","COLD","COLM","COO",
    "COOP","COP","COR","CORT","COST","COTY","CPAY","CPB","CPNG","CPRT","CPT","CR","CRH","CRL","CRM","CROX","CRS","CRUS","CRWD","CSCO","CSGP",
    "CSL","CSX","CTAS","CTRA","CTSH","CTVA","CUBE","CUZ","CVNA","CVS","CVX","CW","CWEN","CWE-A","CXT","CZR","D","DAL","DAR","DASH","DAY","DBX",
    "DCI","DD","DDOG","DDS","DE","DECK","DELL","DG","DGX","DHI","DHR","DINO","DIS","DJT","DKNG","DKS","DLB","DLR","DLTR","DNB","DOC","DOCS",
    "DOCU","DOV","DOW","DOX","DPZ","DRI","DRS","DT","DTE","DTM","DUK","DUOL","DV","DVA","DVN","DXC","DXCM","EA","EBAY","ECG","ECL","ED",
    "EEFT","EFX","EG","EGP","EHC","EIX","EL","ELAN","ELF","ELS","ELV","EME","EMN","EMR","ENPH","ENTG","EOG","EPAM","EPR","EQH","EQIX","EQR",
    "EQT","ES","ESAB","ESI","ESS","ESTC","ETN","ETR","ETSY","EVR","EVRG","EW","EWBC","EXAS","EXC","EXE","EXEL","EXLS","EXP","EXPD","EXPE",
    "EXR","F","FAF","FANG","FAST","FBIN","FCN","FCNCA","FCX","FDS","FDX","FE","FERG","FFIV","FHB","FHN","FI","FICO","FIS","FITB","FIVE",
    "FIX","FLEX","FLO","FLS","FLUT","FMC","FNB","FND","FNF","FOUR","FOX","FOXA","FR","FRHC","FRPT","FRT","FSLR","FTAI","FTI","FTNT","FTV",
    "FWONA","FWONK","FYBR","G","GAP","GD","GDDY","GE","GEHC","GEN","GEV","GFS","GGG","GILD","GIS","GL","GLIBA","GLIBK","GLOB","GLPI","GLW",
    "GM","GME","GMED","GNRC","GNTX","GOOG","GOOGL","GPC","GPK","GPN","GRMN","GS","GTES","GTLB","GTM","GWRE","GWW","GXO","H","HAL","HALO",
    "HAS","HAYW","HBAN","HCA","HD","HEI","HEI-A","HHH","HIG","HII","HIW","HLI","HLNE","HLT","HOG","HOLX","HON","HOOD","HPE","HPQ","HR","HRB",
    "HRL","HSIC","HST","HSY","HUBB","HUBS","HUM","HUN","HWM","HXL","IAC","IBKR","IBM","ICE","IDA","IDXX","IEX","IFF","ILMN","INCY","INFA",
    "INGM","INGR","INSM","INSP","INTC","INTU","INVH","IONS","IOT","IP","IPG","IPGP","IQV","IR","IRDM","IRM","ISRG","IT","ITT","ITW","IVZ",
    "J","JAZZ","JBHT","JBL","JCI","JEF","JHG","JHX","JKHY","JLL","JNJ","JPM","K","KBR","KD","KDP","KEX","KEY","KEYS","KHC","KIM","KKR","KLAC",
    "KMB","KMI","KMPR","KMX","KNSL","KNX","KO","KR","KRC","KRMN","KVUE","L","LAD","LAMR","LAZ","LBRDA","LBRDK","LBTYA","LBTYK","LCID","LDOS",
    "LEA","LECO","LEN","LEN-B","LFUS","LH","LHX","LII","LIN","LINE","LITE","LKQ","LLY","LLYVA","LLYVK","LMT","LNC","LNG","LNT","LNW","LOAR",
    "LOPE","LOW","LPLA","LPX","LRCX","LSCC","LSTR","LULU","LUV","LVS","LW","LYB","LYFT","LYV","M","MA","MAA","MAN","MANH","MAR","MAS","MASI",
    "MAT","MCD","MCHP","MCK","MCO","MDB","MDLZ","MDT","MDU","MEDP","MET","META","MGM","MHK","MIDD","MKC","MKL","MKSI","MKTX","MLI","MLM","MMC",
    "MMM","MNST","MO","MOH","MORN","MOS","MP","MPC","MPW","MPWR","MRK","MRNA","MRP","MRVL","MS","MSA","MSCI","MSFT","MSGS","MSI","MSM","MSTR",
    "MTB","MTCH","MTD","MTDR","MTG","MTN","MTSI","MTZ","MU","MUSA","NBIX","NCLH","NCNO","NDAQ","NDSN","NEE","NEM","NET","NEU","NFG","NFLX","NI",
    "NKE","NLY","NNN","NOC","NOV","NOW","NRG","NSA","NSC","NTAP","NTNX","NTRA","NTRS","NU","NUE","NVDA","NVR","NVST","NVT","NWL","NWS","NWSA",
    "NXST","NYT","O","OC","ODFL","OGE","OGN","OHI","OKE","OKTA","OLED","OLLI","OLN","OMC","OMF","ON","ONON","ONTO","ORCL","ORI","ORLY","OSK",
    "OTIS","OVV","OWL","OXY","OZK","PAG","PANW","PATH","PAYC","PAYX","PB","PCAR","PCG","PCOR","PCTY","PEG","PEGA","PEN","PENN","PEP","PFE","PFG",
    "PFGC","PG","PGR","PH","PHM","PINS","PK","PKG","PLD","PLNT","PLTR","PM","PNC","PNFP","PNR","PNW","PODD","POOL","POST","PPC","PPG","PPL",
    "PR","PRGO","PRI","PRMB","PRU","PSA","PSN","PSTG","PSX","PTC","PVH","PWR","PYPL","QCOM","QGEN","QRVO","QS","QSR","QXO","R","RAL","RARE",
    "RBA","RBC","RBLX","RBRK","RCL","RDDT","REG","REGN","REXR","REYN","RF","RGA","RGEN","RGLD","RH","RHI","RITM","RIVN","RJF","RKLB","RKT",
    "RL","RLI","RMD","RNG","RNR","ROIV","ROK","ROKU","ROL","ROP","ROST","RPM","RPRX","RRC","RRX","RS","RSG","RTX","RVMD","RVTY","RYAN","RYN",
    "S","SAIA","SAIC","SAIL","SAM","SARO","SBAC","SBUX","SCCO","SCHW","SCI","SEB","SEE","SEIC","SF","SFD","SFM","SGI","SHC","SHW","SIRI","SITE",
    "SJM","SKX","SLB","SLGN","SLM","SMCI","SMG","SMMT","SN","SNA","SNDK","SNDR","SNOW","SNPS","SNV","SNX","SO","SOFI","SOLV","SON","SPG","SPGI",
    "SPOT","SPR","SRE","SRPT","SSB","SSD","SSNC","ST","STAG","STE","STLD","STT","STWD","STZ","SUI","SW","SWK","SWKS","SYF","SYK","SYY","T","TAP",
    "TDC","TDG","TDY","TEAM","TECH","TEM","TER","TFC","TFSL","TFX","TGT","THC","THG","THO","TIGO","TJX","TKO","TKR","TLN","TMO","TMUS","TNL","TOL",
    "TOST","TPG","TPL","TPR","TREX","TRGP","TRMB","TROW","TRU","TRV","TSCO","TSLA","TSN","TT","TTC","TTD","TTEK","TTWO","TW","TWLO","TXN","TXRH",
    "TXT","TYL","U","UA","UAA","UAL","UBER","UDR","UGI","UHAL","UHA-B","UHS","UI","ULTA","UNH","UNM","UNP","UPS","URI","USB","USFD","UTHR","UWMC",
    "V","VEEV","VFC","VICI","VIK","VIRT","VKTX","VLO","VLTO","VMC","VMI","VNO","VNOM","VNT","VOYA","VRSK","VRSN","VRT","VRTX","VST","VTR","VTRS",
    "VVV","VZ","W","WAB","WAL","WAT","WBA","WBD","WBS","WCC","WDAY","WDC","WEC","WELL","WEN","WEX","WFC","WFRD","WH","WHR","WING","WLK","WM","WMB",
    "WMS","WMT","WPC","WRB","WSC","WSM","WSO","WST","WTFC","WTM","WTRG","WTW","WU","WWD","WY","WYNN","XEL","XOM","XP","XPO","XRAY","XYL","XYZ","ZTS",
    "TEL","NXPI","YUM","STX","ZBH","ERIE","ZBRA","PSKY"
]

# ─── FECHAS DINÁMICAS ─────────────────────────────────────────────────────────
# Lógica: hoy → fin del mes anterior (t-1) → mismo mes del año anterior (t-12)
# Esto garantiza coherencia sin importar cuándo se corra el script.

today = date.today()

# Último día del mes anterior (t-1): primer día del mes actual menos 1 día
end_date = today.replace(day=1) - relativedelta(days=1)

# Mismo día pero 12 meses atrás (t-12)
start_date = end_date - relativedelta(months=12)

# Para descarga de precios diarios necesitamos un poco más de historia
# (buffer de 5 días hábiles para asegurar que tengamos el cierre de start_date)
fetch_from = start_date - relativedelta(days=7)

print("=" * 60)
print("  SPMO MOMENTUM SCORE CALCULATOR")
print("=" * 60)
print(f"  Fecha de ejecución : {today}")
print(f"  Precio t-12 (inicio): {start_date}  [{start_date.strftime('%b %Y')}]")
print(f"  Precio t-1  (fin)   : {end_date}  [{end_date.strftime('%b %Y')}]")
print(f"  Retornos diarios σ  : {start_date} → {end_date}")
print("=" * 60)

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def get_json(url: str, retries: int = 3) -> any:
    """Descarga y parsea JSON desde una URL con reintentos ante error 429."""
    import urllib.error
    for attempt in range(retries):
        try:
            time.sleep(API_CALL_DELAY)
            response = urlopen(url, cafile=certifi.where())
            data = response.read().decode("utf-8")
            return json.loads(data)
        except urllib.error.HTTPError as e:
            if e.code == 429:
                wait = 10 * (attempt + 1)
                print(f"    ⏳ Rate limit (429) — esperando {wait}s antes de reintentar...")
                time.sleep(wait)
            else:
                raise
    raise RuntimeError(f"Fallo tras {retries} intentos por rate limit (429)")


def fetch_prices(ticker: str) -> pd.DataFrame:
    """
    Descarga precios diarios EOD desde FMP y retorna un DataFrame
    con columnas [date, close] ordenado de más antiguo a más reciente.
    """
    url = (
        f"https://financialmodelingprep.com/stable/historical-price-eod/light"
        f"?symbol={ticker}"
        f"&from={fetch_from}"
        f"&to={end_date}"
        f"&apikey={API_KEY}"
    )
    raw = get_json(url)

    if not raw:
        return pd.DataFrame()

    df = pd.DataFrame(raw)
    df["date"] = pd.to_datetime(df["date"])
    df = df[["date", "price"]].rename(columns={"price": "close"}).sort_values("date").reset_index(drop=True)
    return df


def fetch_market_cap(ticker: str) -> float:
    """Obtiene la capitalización de mercado actual desde FMP."""
    url = (
        f"https://financialmodelingprep.com/stable/market-capitalization"
        f"?symbol={ticker}"
        f"&apikey={API_KEY}"
    )
    raw = get_json(url)

    if not raw:
        return np.nan

    # El endpoint devuelve una lista; tomamos el primer registro
    record = raw[0] if isinstance(raw, list) else raw
    return float(record.get("marketCap", np.nan))


def fetch_beta(ticker: str) -> float:
    """Obtiene el beta del ticker desde el endpoint /profile de FMP."""
    url = (
        f"https://financialmodelingprep.com/stable/profile"
        f"?symbol={ticker}"
        f"&apikey={API_KEY}"
    )
    raw = get_json(url)
    if not raw:
        return np.nan
    record = raw[0] if isinstance(raw, list) else raw
    return float(record.get("beta", np.nan))


def fetch_prices_full(ticker: str, days_back: int = 90) -> pd.DataFrame:
    """
    Descarga precios diarios EOD usando el endpoint /light (price + volume).
    Aproxima high y low como max/min entre close_hoy y close_ayer,
    ya que el endpoint /light no expone high/low reales.
    Suficiente para calcular ATR basado en gaps de cierre.
    """
    from_date = today - relativedelta(days=days_back)
    url = (
        f"https://financialmodelingprep.com/stable/historical-price-eod/light"
        f"?symbol={ticker}"
        f"&from={from_date}"
        f"&to={today}"
        f"&apikey={API_KEY}"
    )
    raw = get_json(url)

    if not raw:
        return pd.DataFrame()

    df = pd.DataFrame(raw)
    df["date"] = pd.to_datetime(df["date"])
    df = df[["date", "price"]].rename(columns={"price": "close"})
    df = df.sort_values("date").reset_index(drop=True)

    # Aproximar high/low a partir de cierres consecutivos
    df["prev_close"] = df["close"].shift(1)
    df["high"] = df[["close", "prev_close"]].max(axis=1)
    df["low"]  = df[["close", "prev_close"]].min(axis=1)
    df = df.drop(columns=["prev_close"])

    return df


def calc_sma(df: pd.DataFrame, period: int = 50) -> float:
    """Calcula la SMA de los últimos `period` cierres disponibles."""
    closes = df["close"].dropna()
    if len(closes) < period:
        return np.nan
    return closes.iloc[-period:].mean()


def calc_atr(df: pd.DataFrame, period: int = 15) -> float:
    """
    Calcula el ATR con suavizado RMA (Wilder), igual a TradingView.
    Necesita columnas: high, low, close.
    """
    if len(df) < period + 1:
        return np.nan

    df = df.copy().reset_index(drop=True)

    # True Range
    df["prev_close"] = df["close"].shift(1)
    df["tr"] = df.apply(
        lambda r: max(
            r["high"] - r["low"],
            abs(r["high"] - r["prev_close"]) if not np.isnan(r["prev_close"]) else 0,
            abs(r["low"]  - r["prev_close"]) if not np.isnan(r["prev_close"]) else 0,
        ),
        axis=1,
    )

    tr = df["tr"].dropna().values

    if len(tr) < period:
        return np.nan

    # Seed: SMA de los primeros `period` valores
    atr = tr[:period].mean()

    # RMA (Wilder smoothing)
    for val in tr[period:]:
        atr = (atr * (period - 1) + val) / period

    return atr


def check_entry_signal(ticker: str) -> dict:
    """
    Descarga datos recientes y evalúa las condiciones de entrada:
      C1: precio_actual <= SMA50
      C2: precio_actual - ATR15 <= SMA50
    Retorna un dict con los valores calculados y el estado de la señal.
    """
    df = fetch_prices_full(ticker, days_back=90)

    if df.empty or len(df) < 51:
        return {"Ticker": ticker, "Precio": np.nan, "SMA50": np.nan,
                "ATR15": np.nan, "P_minus_ATR": np.nan,
                "C1 (P<=SMA)": "FALSE", "C2 (P-ATR<=SMA)": "FALSE", "Señal": "Sin datos"}

    precio  = df["close"].iloc[-1]
    sma50   = calc_sma(df, period=50)
    atr15   = calc_atr(df, period=15)

    if np.isnan(sma50) or np.isnan(atr15):
        return {"Ticker": ticker, "Precio": precio, "SMA50": sma50,
                "ATR15": atr15, "P_minus_ATR": np.nan,
                "C1 (P<=SMA)": "FALSE", "C2 (P-ATR<=SMA)": "FALSE", "Señal": "Calculo incompleto"}

    p_minus_atr = precio - atr15
    c1 = precio       <= sma50
    c2 = p_minus_atr  <= sma50
    activa = c1 and c2

    return {
        "Ticker"     : ticker,
        "Precio"     : round(precio, 2),
        "SMA50"      : round(sma50, 2),
        "ATR15"      : round(atr15, 2),
        "P_minus_ATR": round(p_minus_atr, 2),
        "C1 (P<=SMA)": "TRUE" if c1 else "FALSE",
        "C2 (P-ATR<=SMA)": "TRUE" if c2 else "FALSE",
        "Señal"      : "ENTRADA" if activa else "Sin señal",
    }


def get_monthly_close(df: pd.DataFrame, target_date: date) -> float:
    """
    Retorna el precio de cierre más cercano (igual o anterior) a target_date.
    Útil para obtener el cierre de fin de mes aunque no sea día hábil.
    """
    mask = df["date"] <= pd.Timestamp(target_date)
    subset = df[mask]
    if subset.empty:
        return np.nan
    return subset.iloc[-1]["close"]


# ─── EXPORTACIÓN A EXCEL ──────────────────────────────────────────────────────

def style_header(ws, row: int, col_count: int, fill_hex: str):
    """Aplica estilo de encabezado a una fila."""
    fill = PatternFill("solid", start_color=fill_hex, end_color=fill_hex)
    font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    border_side = Side(style="thin", color="CCCCCC")
    border = Border(bottom=border_side)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def style_data_row(ws, row: int, col_count: int, alt: bool):
    """Aplica estilo alternado a filas de datos."""
    fill_hex = "F2F2F2" if alt else "FFFFFF"
    fill = PatternFill("solid", start_color=fill_hex, end_color=fill_hex)
    font = Font(name="Arial", size=10)
    border_side = Side(style="thin", color="DDDDDD")
    border = Border(bottom=border_side)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def write_df_to_sheet(ws, df: pd.DataFrame, title: str, header_color: str):
    """Escribe un DataFrame en una hoja con título y formato."""
    # Título
    ws.append([title])
    title_cell = ws.cell(row=ws.max_row, column=1)
    title_cell.font = Font(bold=True, name="Arial", size=12, color="1F3864")
    title_cell.alignment = Alignment(horizontal="left")
    ws.append([f"Generado: {date.today()}  |  Período: {start_date} → {end_date}"])
    ws.cell(row=ws.max_row, column=1).font = Font(name="Arial", size=9, italic=True, color="666666")
    ws.append([])  # fila vacía

    cols = list(df.columns)
    header_row = ws.max_row + 1
    ws.append(cols)
    style_header(ws, header_row, len(cols), header_color)

    for i, row_data in enumerate(df.itertuples(index=False), start=1):
        ws.append(list(row_data))
        style_data_row(ws, ws.max_row, len(cols), alt=(i % 2 == 0))

    # Ancho automático de columnas
    for col_idx, col_name in enumerate(cols, start=1):
        col_values = [len(str(v)) for v in df.iloc[:, col_idx - 1]] if not df.empty else []
        max_len = max([len(str(col_name))] + col_values)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)


def export_to_excel(df_momentum: pd.DataFrame, df_active: pd.DataFrame,
                    df_inactive: pd.DataFrame, path: str):
    """Crea el archivo Excel con tres hojas separadas."""
    os.makedirs(os.path.dirname(path), exist_ok=True)

    wb = Workbook()

    # ── Hoja 1: Momentum Scores ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Momentum Scores"
    ws1.sheet_view.showGridLines = False
    ws1.row_dimensions[1].height = 5

    display = df_momentum[[
        "Ticker", "R_12_1_pct", "Sigma_diaria", "Score_bruto", "Momentum_Score", "Peso_pct"
    ]].rename(columns={
        "R_12_1_pct"    : "Retorno 12m (%)",
        "Sigma_diaria"  : "Sigma diaria",
        "Score_bruto"   : "Score bruto",
        "Momentum_Score": "Score Z",
        "Peso_pct"      : "Peso ETF (%)",
    }).reset_index(drop=True)
    display.insert(0, "Rank", range(1, len(display) + 1))
    write_df_to_sheet(ws1, display, "MOMENTUM SCORES — S&P 500 Momentum Index", "1F3864")

    # ── Hoja 2: Señales Activas ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Senales Activas")
    ws2.sheet_view.showGridLines = False
    ws2.row_dimensions[1].height = 5
    write_df_to_sheet(ws2, df_active, "SEÑALES DE ENTRADA ACTIVAS (C1 y C2 cumplidas)", "1A7A4A")

    # ── Hoja 3: Sin Señal ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Sin Senal")
    ws3.sheet_view.showGridLines = False
    ws3.row_dimensions[1].height = 5
    write_df_to_sheet(ws3, df_inactive, "SIN SEÑAL — Top 50 restantes", "555555")

    wb.save(path)
    print(f"\n  ✅ Excel guardado en: {path}")


# ─── CÁLCULO PRINCIPAL ────────────────────────────────────────────────────────

results = []

for ticker in TICKERS:
    print(f"\n  Procesando: {ticker}")

    # 1. Descargar precios
    df_prices = fetch_prices(ticker)

    if df_prices.empty:
        print(f"    ⚠️  Sin datos de precios para {ticker}")
        continue

    # 2. Precio t-12 y t-1 para el retorno simple R_12,1
    price_t12 = get_monthly_close(df_prices, start_date)
    price_t1  = get_monthly_close(df_prices, end_date)

    if np.isnan(price_t12) or np.isnan(price_t1) or price_t12 == 0:
        print(f"    ⚠️  Precios insuficientes para {ticker}")
        continue

    R_12_1 = (price_t1 / price_t12) - 1

    # 3. Volatilidad σ: desviación estándar de retornos diarios en el período
    period_mask = (
        (df_prices["date"] >= pd.Timestamp(start_date)) &
        (df_prices["date"] <= pd.Timestamp(end_date))
    )
    df_period = df_prices[period_mask].copy()

    if len(df_period) < 30:
        print(f"    ⚠️  Pocos días hábiles para calcular σ en {ticker}")
        continue

    df_period["daily_return"] = df_period["close"].pct_change()
    sigma = df_period["daily_return"].std()

    if sigma == 0 or np.isnan(sigma):
        print(f"    ⚠️  Volatilidad inválida para {ticker}")
        continue

    # 4. Momentum Score bruto = R_12,1 / σ
    raw_score = R_12_1 / sigma

    # 5. Market Cap
    mkt_cap = fetch_market_cap(ticker)

    print(f"    P(t-12): ${price_t12:>10.2f}  |  P(t-1): ${price_t1:>10.2f}")
    print(f"    R_12,1 : {R_12_1:>+.4f} ({R_12_1*100:+.2f}%)")
    print(f"    σ diaria: {sigma:.6f}  |  Score bruto: {raw_score:.4f}")
    print(f"    Mkt Cap: ${mkt_cap:,.0f}" if not np.isnan(mkt_cap) else "    Mkt Cap: N/A")

    results.append({
        "Ticker"       : ticker,
        "P_t12"        : round(price_t12, 2),
        "P_t1"         : round(price_t1, 2),
        "R_12_1"       : round(R_12_1, 6),
        "R_12_1_pct"   : round(R_12_1 * 100, 2),
        "Sigma_diaria" : round(sigma, 6),
        "Score_bruto"  : round(raw_score, 4),
        "Mkt_Cap"      : mkt_cap,
    })

# ─── ESTANDARIZACIÓN (Z-SCORE) ────────────────────────────────────────────────

if not results:
    print("\n  ❌ No se obtuvieron resultados. Verifica tu API key y tickers.")
else:
    df_result = pd.DataFrame(results)

    scores = df_result["Score_bruto"]
    df_result["Momentum_Score"] = (scores - scores.mean()) / scores.std()

    # Peso proporcional = Mkt_Cap × Momentum_Score (solo scores positivos)
    df_result["peso_raw"] = df_result.apply(
        lambda r: r["Mkt_Cap"] * r["Momentum_Score"]
        if (r["Momentum_Score"] > 0 and not np.isnan(r["Mkt_Cap"]))
        else 0,
        axis=1
    )
    total_peso = df_result["peso_raw"].sum()
    df_result["Peso_pct"] = (
        (df_result["peso_raw"] / total_peso * 100).round(2)
        if total_peso > 0 else 0
    )

    # Ordenar de mayor a menor Momentum Score
    df_result = df_result.sort_values("Momentum_Score", ascending=False).reset_index(drop=True)
    df_result.index += 1  # ranking desde 1

    # ─── OUTPUT FINAL ─────────────────────────────────────────────────────────

    print("\n")
    print("=" * 60)
    print("  RESULTADOS FINALES — ORDENADOS POR MOMENTUM SCORE")
    print("=" * 60)

    display_cols = [
        "Ticker", "R_12_1_pct", "Sigma_diaria",
        "Score_bruto", "Momentum_Score", "Peso_pct"
    ]

    pd.set_option("display.float_format", "{:.4f}".format)
    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", 120)

    print(df_result[display_cols].rename(columns={
        "R_12_1_pct"   : "Retorno 12m (%)",
        "Sigma_diaria" : "Sigma diaria",
        "Score_bruto"  : "Score bruto",
        "Momentum_Score": "Score Z",
        "Peso_pct"     : "Peso ETF (%)",
    }).to_string())

    print("\n")
    print("=" * 60)
    print(f"  Período analizado: {start_date} → {end_date}")
    print(f"  Tickers procesados: {len(df_result)} / {len(TICKERS)}")
    print("=" * 60)

    # ─── PARTE 2: SEÑAL DE ENTRADA — TOP 50 POR SCORE BRUTO ──────────────────

    # Ordenar por Score bruto y tomar top 50
    top50 = (
        df_result.sort_values("Score_bruto", ascending=False)
        .head(50)["Ticker"]
        .tolist()
    )

    print("\n")
    print("=" * 60)
    print("  DETECCIÓN DE SEÑAL DE ENTRADA — TOP 50 (Score bruto)")
    print("  Condiciones:")
    print("    C1: Precio actual  <= SMA 50")
    print("    C2: Precio - ATR15 <= SMA 50")
    print("=" * 60)

    signal_results = []

    for ticker in top50:
        print(f"  Analizando señal: {ticker}", end=" ... ", flush=True)
        sig = check_entry_signal(ticker)
        signal_results.append(sig)
        print(f"  → {sig['Señal']}")

    df_signals = pd.DataFrame(signal_results)

    # Separar con señal activa vs sin señal
    df_active  = df_signals[df_signals["Señal"] == "ENTRADA"].reset_index(drop=True)
    df_inactive = df_signals[df_signals["Señal"] != "ENTRADA"].reset_index(drop=True)

    # Obtener Beta solo para tickers con señal activa
    if not df_active.empty:
        print("\n  Obteniendo Beta para tickers con señal activa...")
        betas = []
        for ticker in df_active["Ticker"]:
            b = fetch_beta(ticker)
            betas.append(round(b, 4) if not np.isnan(b) else np.nan)
            print(f"    {ticker}: Beta = {betas[-1]}")
        df_active.insert(df_active.columns.get_loc("Señal"), "Beta", betas)

    df_active.index  += 1
    df_inactive.index += 1

    print("\n")
    print("─" * 60)
    print(f"  🟢 TICKERS CON SEÑAL DE ENTRADA ACTIVA: {len(df_active)}")
    print("─" * 60)

    if df_active.empty:
        print("  Ningún ticker del Top 50 cumple ambas condiciones hoy.")
    else:
        pd.set_option("display.float_format", "{:.2f}".format)
        print(df_active.to_string())

    print("\n")
    print("─" * 60)
    print(f"  ⚪ SIN SEÑAL ({len(df_inactive)} tickers)")
    print("─" * 60)
    print(df_inactive[["Ticker", "Precio", "SMA50", "ATR15",
                        "C1 (P<=SMA)", "C2 (P-ATR<=SMA)"]].to_string())

    print("\n")
    print("=" * 60)
    print(f"  Fecha de análisis  : {today}")
    print(f"  Top 50 evaluados   : {len(top50)}")
    print(f"  Con señal activa   : {len(df_active)}")
    print("=" * 60)

    # ─── EXPORTAR A EXCEL ─────────────────────────────────────────────────────
    export_to_excel(df_result, df_active, df_inactive, EXCEL_PATH)